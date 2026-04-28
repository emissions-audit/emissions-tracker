"""EU Emissions Trading System (EU ETS) data source.

Downloads verified emissions data from the European Commission's
Transaction Log (EUTL) Excel exports and maps installations to
company tickers.

EU ETS reports emissions at installation level in **tonnes CO2e**
(t_co2e — Scope 1 only, since the scheme covers direct emissions
from regulated installations).

PARTIAL COVERAGE NOTE: EU ETS only covers EU-located installations.
For non-EU-headquartered companies (e.g. XOM, CVX), this source will
report substantially less than the company's global Scope 1. This is
correct behavior for cross-validation; CV engine semantics for partial
sources are tracked in ET-85.
"""

import asyncio
import io
import sys

import httpx

# Realistic browser UA — the EC's climate.ec.europa.eu CDN 429s bare httpx/python-
# requests clients on first hit.
_EU_ETS_HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "application/octet-stream,*/*;q=0.8"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# Polite pause between per-year workbook downloads. The EC rate-limits bursts
# from a single IP (common on shared CI runners).
_EU_ETS_INTER_REQUEST_DELAY_S = 2.0

# Retry 429s a few times with a linear backoff before giving up on a year.
_EU_ETS_MAX_RETRIES = 3
_EU_ETS_RETRY_BACKOFF_S = 5.0

from src.pipeline.company_mapping import resolve_ticker
from src.pipeline.sources.base import BaseSource, RawEmission

# The EUTL publishes compliance data as annual Excel workbooks with UUID-based
# download URLs (no computable slug pattern). Each year's workbook contains the
# VERIFIED_EMISSIONS_{year} column for that year only, so we download each one
# separately. Refresh this map when the EC publishes new years — the canonical
# list lives at
# https://climate.ec.europa.eu/eu-action/eu-emissions-trading-system-eu-ets/union-registry_en
EU_ETS_COMPLIANCE_URLS = {
    2020: "https://climate.ec.europa.eu/document/download/2d3e055d-ba0e-46db-b4c1-e3a9210d7ddb_en?filename=compliance_2020_code_en.xlsx",
    2021: "https://climate.ec.europa.eu/document/download/86a31a71-dff3-4729-86d0-943685c20dc1_en?filename=compliance_2021_code_en.xlsx",
    2022: "https://climate.ec.europa.eu/document/download/7e7268a1-fa21-4f73-b368-6e9571262e2f_en?filename=compliance_2022_code_en.xlsx",
    2023: "https://climate.ec.europa.eu/document/download/42495a32-cb4c-4772-9a2a-d08781c8ed61_en?filename=compliance_2023_code_en.xlsx",
    2024: "https://climate.ec.europa.eu/document/download/b80300cf-7608-405d-969e-8b016687640e_en?filename=compliance_2024_code_en.xlsx",
}

EU_ETS_LATEST_AVAILABLE_YEAR = max(EU_ETS_COMPLIANCE_URLS)


def _resolve_installation_ticker(installation_name: str) -> str:
    """Resolve an EU ETS installation name to a stock ticker.

    EU ETS installation names typically follow the pattern
    ``"Company Name - Facility Description"``.  We try the full name first,
    then progressively shorter prefixes (splitting on `` - ``), and finally
    each whitespace-delimited token.  If nothing matches, the raw
    installation name is returned as-is.
    """
    # 1. Try the full installation name
    ticker = resolve_ticker(installation_name)
    if ticker is not None:
        return ticker

    # 2. Try prefix segments (split on " - ")
    parts = installation_name.split(" - ")
    if len(parts) > 1:
        # Try progressively shorter prefixes: "A - B - C" -> "A - B", then "A"
        for i in range(len(parts) - 1, 0, -1):
            prefix = " - ".join(parts[:i])
            ticker = resolve_ticker(prefix)
            if ticker is not None:
                return ticker

    # 3. Try the first segment alone (the company-name portion)
    first_segment = parts[0].strip()
    if first_segment != installation_name:
        ticker = resolve_ticker(first_segment)
        if ticker is not None:
            return ticker

    # 4. Try progressively shorter word prefixes of the first segment.
    #    E.g. "Shell Deutschland Oil GmbH" -> "Shell Deutschland Oil",
    #    "Shell Deutschland", "Shell"
    words = first_segment.split()
    for end in range(len(words) - 1, 0, -1):
        candidate = " ".join(words[:end])
        ticker = resolve_ticker(candidate)
        if ticker is not None:
            return ticker

    # 5. Fall back to the raw installation name
    return installation_name


def _resolve_operator_name(row: dict, installations_lookup: dict[str, str] | None = None) -> str:
    """Extract the best available operator name from a compliance workbook row.

    2023+ EC workbooks changed INSTALLATION_NAME to a numeric code.  This
    function checks multiple columns in priority order and falls back to
    an optional installations lookup dict (installation_id -> operator_name).
    """
    # Priority 1: ACCOUNT_HOLDER_NAME (present in some workbook versions)
    account_holder = row.get("ACCOUNT_HOLDER_NAME") or row.get("OPERATOR_NAME")
    if account_holder and isinstance(account_holder, str) and not account_holder.isdigit():
        return account_holder.strip()

    # Priority 2: INSTALLATION_NAME (reliable pre-2023)
    installation_name = str(row.get("INSTALLATION_NAME", "")).strip()
    if installation_name and not installation_name.isdigit():
        return installation_name

    # Priority 3: Installations lookup by INSTALLATION_IDENTIFIER or numeric code
    if installations_lookup:
        inst_id = str(row.get("INSTALLATION_IDENTIFIER", "")).strip()
        if inst_id and inst_id in installations_lookup:
            return installations_lookup[inst_id]
        if installation_name and installation_name in installations_lookup:
            return installations_lookup[installation_name]

    # Fallback: return whatever we have (may be numeric — won't resolve to a ticker)
    return installation_name or ""


def parse_eu_ets_data(
    records: list[dict],
    years: list[int],
    installations_lookup: dict[str, str] | None = None,
) -> list[RawEmission]:
    """Parse pre-extracted EU ETS installation dicts into RawEmission records.

    Each *record* represents one installation row from the EUTL Excel file
    (already converted to a flat dict).  The EC publishes one workbook per
    compliance year.  The emissions column is either
    ``TOTAL_VERIFIED_EMISSIONS`` (current format) or
    ``VERIFIED_EMISSIONS_{year}`` (legacy multi-year format).

    Parameters
    ----------
    records:
        List of dicts from a single year's compliance workbook.
    years:
        Which reporting years to extract (typically a single-element list
        matching the workbook year).
    installations_lookup:
        Optional mapping from installation ID/code to operator name,
        for resolving 2023+ numeric INSTALLATION_NAME values.

    Returns
    -------
    list[RawEmission]
        One entry per (installation, year) pair where the value is non-None.
    """
    results: list[RawEmission] = []

    for row in records:
        operator_name = _resolve_operator_name(row, installations_lookup)

        ticker = _resolve_installation_ticker(operator_name)

        for year in years:
            # Current EC format: single TOTAL_VERIFIED_EMISSIONS column per workbook.
            # Legacy format: VERIFIED_EMISSIONS_{year} columns.
            value = row.get("TOTAL_VERIFIED_EMISSIONS")
            if value is None:
                value = row.get(f"VERIFIED_EMISSIONS_{year}")
            if value is None:
                continue

            # Excel cells may arrive as strings — coerce to float.
            try:
                value = float(value)
            except (TypeError, ValueError):
                continue

            source_url = (
                f"https://climate.ec.europa.eu/eu-action/eu-emissions-trading-system/"
                f"union-registry_en#tab-compliance-data"
            )

            results.append(
                RawEmission(
                    company_ticker=ticker,
                    year=year,
                    scope="Scope 1",
                    value=value,
                    unit="t_co2e",
                    methodology="eu_ets_verified",
                    verified=True,
                    source_url=source_url,
                    filing_type="eu_ets",
                    parser_used="excel",
                )
            )

    return results


class EuEtsSource(BaseSource):
    """EU ETS verified emissions data source."""

    name = "eu_ets"

    async def fetch_emissions(
        self, tickers: list[str], years: list[int]
    ) -> list[RawEmission]:
        """Download per-year EU ETS Excel workbooks and parse installation emissions.

        Each year has its own workbook (UUID-based URLs — see
        :data:`EU_ETS_COMPLIANCE_URLS`).  For each requested year that has a
        known URL, fetch the workbook, parse the ``VERIFIED_EMISSIONS_{year}``
        column, and map installation names to tickers. Silently skip years we
        don't have URLs for (e.g. current or future year where the EC has not
        yet published).

        For 2023+ workbooks where INSTALLATION_NAME is numeric, the parser
        falls back to ACCOUNT_HOLDER_NAME or an installations lookup dict.

        Returns a merged list across all downloaded years.
        """
        results: list[RawEmission] = []
        target_years = [y for y in years if y in EU_ETS_COMPLIANCE_URLS]

        # Build installations lookup if any target year is 2023+
        installations_lookup = None
        if any(y >= 2023 for y in target_years):
            installations_lookup = await self._load_installations_lookup()

        for i, year in enumerate(target_years):
            if i > 0:
                await asyncio.sleep(_EU_ETS_INTER_REQUEST_DELAY_S)
            try:
                records = await self._download_and_parse(year)
            except httpx.HTTPError as e:
                print(f"  eu_ets {year}: DOWNLOAD FAILED — {type(e).__name__}: {e}")
                continue
            except Exception as e:
                print(f"  eu_ets {year}: PARSE FAILED — {type(e).__name__}: {e}")
                continue
            parsed = parse_eu_ets_data(records, [year], installations_lookup)
            print(f"  eu_ets {year}: {len(records)} rows → {len(parsed)} emissions")
            results.extend(parsed)

        if tickers:
            ticker_set = {t.upper() for t in tickers}
            results = [r for r in results if r.company_ticker.upper() in ticker_set]

        return results

    async def _load_installations_lookup(self) -> dict[str, str]:
        """Load installation ID → operator name mapping.

        Checks for a local data/eu_ets_installations.csv first. If not found,
        returns an empty dict (graceful degradation — ACCOUNT_HOLDER_NAME in
        the workbook is the primary fallback for 2023+ data).
        """
        import csv
        from pathlib import Path

        local_path = Path(__file__).resolve().parents[3] / "data" / "eu_ets_installations.csv"
        if local_path.exists():
            lookup = {}
            with open(local_path, encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    inst_id = row.get("INSTALLATION_IDENTIFIER", "").strip()
                    inst_code = row.get("INSTALLATION_NAME", "").strip()
                    operator = row.get("ACCOUNT_HOLDER_NAME", "").strip()
                    if operator:
                        if inst_id:
                            lookup[inst_id] = operator
                        if inst_code:
                            lookup[inst_code] = operator
            if lookup:
                print(f"  eu_ets: loaded {len(lookup)} installation lookups from {local_path}")
            return lookup

        print(
            "  eu_ets: no data/eu_ets_installations.csv found — "
            "relying on ACCOUNT_HOLDER_NAME column for 2023+ workbooks",
            file=sys.stderr,
        )
        return {}

    async def _download_and_parse(self, year: int) -> list[dict]:
        """Download a single year's Excel workbook and convert rows to dicts."""
        from openpyxl import load_workbook

        url = EU_ETS_COMPLIANCE_URLS[year]

        async with httpx.AsyncClient(
            timeout=120.0, headers=_EU_ETS_HTTP_HEADERS
        ) as client:
            for attempt in range(_EU_ETS_MAX_RETRIES):
                response = await client.get(url, follow_redirects=True)
                if response.status_code != 429:
                    break
                print(f"  eu_ets {year}: 429 rate-limited (attempt {attempt + 1}/{_EU_ETS_MAX_RETRIES})")
                if attempt < _EU_ETS_MAX_RETRIES - 1:
                    await asyncio.sleep(_EU_ETS_RETRY_BACKOFF_S * (attempt + 1))
            response.raise_for_status()

        wb = load_workbook(filename=io.BytesIO(response.content), read_only=True)
        ws = wb.active

        # Auto-detect header row by scanning for VERIFIED_EMISSIONS or REGISTRY_CODE.
        all_rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, row in enumerate(all_rows):
            cell_strs = [str(c).strip() if c else "" for c in row]
            if any("VERIFIED_EMISSIONS" in c for c in cell_strs) or "REGISTRY_CODE" in cell_strs:
                header_idx = i
                break

        if header_idx is None:
            print(f"  eu_ets {year}: no header row found in {len(all_rows)} rows")
            wb.close()
            return []

        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(all_rows[header_idx])]
        data_rows = all_rows[header_idx + 1:]

        records: list[dict] = []
        for row in data_rows:
            record = dict(zip(headers, row))
            records.append(record)

        wb.close()
        return records
